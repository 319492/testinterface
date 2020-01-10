# -*- coding: utf-8 -*-
# @Time    : 2020/1/6 9:08 上午
# @Author  : Silver
# @File    : testread_excel.py
# @Project_name:
import os


import openpyxl


class Read_Excle():
    def read_work(self,case_path):
        return openpyxl.load_workbook(case_path)

#获取sheet页
    def get_data(self,sheet_name,case_path):
        read_sheet = self.read_work(case_path)[sheet_name]
        return read_sheet



    #获取sheet页的所有行read_sheet
    def get_excel_row(self,sheet_name,case_path):
        get_row = self.get_data(sheet_name,case_path).max_row
        return get_row

    # 获取sheet页的单元格
    def get_excel(self, sheet_name, row, col,case_path):
        get_excel = self.get_data(sheet_name,case_path).cell(row, col)
        return get_excel

    #获取sheet页的单元格-参数
    def get_excel_data(self,sheet_name,row,col,case_path):
        get_excel_data = self.get_excel(sheet_name,row,col,case_path).value
        return get_excel_data

if __name__ == '__main__':
    a = Read_Excle()
    a.get_excel("test",4,4,"../data/接口测试用例.xlsx")
    print()

    #写入excel值1
    # def write_excel(self,sheet_name,case_path,row,col,value):
    #     wb = openpyxl.Workbook()
    #     ws = wb.active
    #     ws.title = sheet_name
    #     ws.cell(row, col, value)
    #     wb.save(case_path)

#
# if __name__ == '__main__':
#     a = Read_Excle()
#     aa=a.write_excel("Sheet1","../data/登陆接口测试用例.xlsx",4,9,"ceshitongguo ")
#     print(aa)



#     # #写入excel方法2
    # # def WriteData(self, case_path,row,col,data):
    # #     sheetName = xlrd.open_workbook(case_path)
    # #     # a = sheetName.sheet_by_name(sheet_name).nrows
    # #     wb = copy(sheetName)
    # #     sheet = wb.get_sheet(0)
    # #     sheet.write(row, col, data)
    # #     os.remove(case_path)
    # #     wb.save(case_path)
    # #引用WorkBook直接写入excel
    # # def write_ecxel_workbook(self,sheet_name,cell,data,case_path):
    # #     # 创建Workbook，并默认会创建一个空表,名称为：Sheet
    # #     wb = Workbook()
    # #     # 获取默认的sheet
    # #     ws1 = wb.active
    # #     # 设置Sheet名称
    # #     ws1.title = sheet_name
    # #     # 写入单个单元格
    # #     ws1[cell] = data
    # #     # 保存
    # #     wb.save(case_path)
    #
    #








# if __name__ == '__main__':
#     a = Read_Excle()
#     a.WriteData("../data/登陆接口测试用例.xlsx",1,13787878887)











# from read_excel import Excel_Const
# a = Read_Excle()
# read_a = a.get_excel_row("Sheet1")
# #读取用例中所有id的值
# with open("data_config.yaml") as fp:
#     data_yaml = yaml.load(fp)
# for i in range(2,read_a):
#     print(a.get_excel_data("Sheet1",i,int(Excel_Const.get_case_input())))


# res=a.get_excel_data("Sheet1",2,int(Excel_Const.get_case_id()))
# print(res)

# read_work=openpyxl.load_workbook("登陆接口测试用例.xlsx")
# read_sheet = read_work["Sheet1"]
# print(read_sheet["F3"].value)

# from read_excel import excelcolum
#
# rd = Read_excel()
# get_count=rd.get_excel_count("工作表1")
# for i in range(2,get_count):
#     print(rd.get_cell_value("工作表1",i,int(excelcolum.get_case_url())))