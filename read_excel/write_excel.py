# -*- coding: utf-8 -*-
# @Time    : 2020/1/9 9:01 上午
# @Author  : Silver
# @File    : write_excel.py
# @Project_name:
import openpyxl
from openpyxl import Workbook
class writeExcel():
    #excel新建文件写入
    def write_ecxel_workbook(self, sheet_name,row,col, data, case_path):
        # 创建Workbook，并默认会创建一个空表,
        wb = Workbook()
        # 获取默认的sheet
        ws1 = wb.active
        # 设置Sheet名称
        ws1.title = sheet_name
        # 写入单个单元格
        ws1.cell(row,col,data)
        # 保存
        wb.save(case_path)
        print ("保存成功")
    #excel追加写入文件
    def write_old_excel(self,case_path,row,col,value):
        data = openpyxl.load_workbook(case_path)
        #获取第一张表
        sheetnames = data.get_sheet_names()
        table = data.get_sheet_by_name(sheetnames[0])
        #指定单元格赋值
        table.cell(row,col).value = value
        #保存文件
        data.save(case_path)
# if __name__ == '__main__':
#     a = writeExcel()
#     a.write_old_excel("../data/登陆接口测试用例.xlsx",3,3,"ds")






# if __name__ == '__main__':
#     wr = writeExcel()
#     wr.write_ecxel_workbook("test","D2","成功","../data/接口测试用例.xlsx")

#
#
# #创建Workbook，并默认会创建一个空表,名称为：Sheet
# wb = Workbook()
# #获取默认的sheet
# ws1 = wb.active
# #设置Sheet名称
# ws1.title = 'Sheet1'
# #写入单个单元格
# # ws1['A1'] = '标题列1'
# # #写入多个单元格(从有数据的行的下一行写入)
# ws1.append(['张三'])
#
# #创建一个新sheet，可以指定名称
# # ws2 = wb.create_sheet('Sheet2')
#
# #复制Sheet1，新sheet名称为Sheet1 Copy
# # ws3 = wb.copy_worksheet(wb['Sheet1'])
#
# #打印所有表名
# print(wb.sheetnames)
#
# #保存
# wb.save('../data/接口测试用例.xlsx')